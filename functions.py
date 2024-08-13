import os
import pandas as pd
import math
from tripLeaderManager import TripLeaderManager, TripLeader
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill


def get_user_input(header_name):
    return input(f"Please enter the excel cell for the '{header_name}' header: ")


def excel_to_df_indices(cell_reference):
    # Extract the column letter and row number from the Excel cell reference
    column_letter = cell_reference[0].upper()
    row_number = int(cell_reference[1:])

    # Convert column letter to zero-based index (e.g., "A" -> 0, "B" -> 1)
    colIndex = ord(column_letter) - ord("A")

    # Convert row number to zero-based index (e.g., 1 -> 0, 2 -> 1)
    # remove 2 to account for the empty row at the top being automatically removed
    rowIndex = row_number - 2

    return rowIndex, colIndex


def excel_to_df_cell(dates_cell, trip_cell, preferences_cell, name_cell):
    dates_index = excel_to_df_indices(dates_cell)
    trip_index = excel_to_df_indices(trip_cell)
    preferences_index = excel_to_df_indices(preferences_cell)
    name_index = excel_to_df_indices(name_cell)

    return dates_index, trip_index, preferences_index, name_index


def get_sheet_names(sheet_name):
    return (
        int(
            input(
                f"Please enter the index of the sheet for the '{sheet_name}' sheet (1st sheet is 1): "
            )
        )
        - 1
    )


def createLeader(
    prefsDF,
    tripLeaderDF,
    numberOfTrips,
    trip_leader_manager,
    nameXY,
    prefXY,
    tripXY,
    file_path,
):
    # leader1 = TripLeader("John Doe", [10, 3, 5, 8, 2, 7, 1, 9, 6, 4])
    # manager.add_trip_leader(leader1)
    name = tripLeaderDF.iloc[nameXY[0], nameXY[1]]

    currentRow = prefXY[0] + 1

    prefs = []

    # while the corresponding trip is not empty, add the preferences to the list
    while currentRow < len(prefsDF) and not pd.isnull(
        prefsDF.iloc[currentRow, tripXY[1]]
    ):
        prefs.append(prefsDF.iloc[currentRow, prefXY[1]])
        currentRow += 1

    # if the name is a float, it means it is NaN, and we should print that there is an error, but not raise an exception
    if isinstance(name, float):
        print(f"Error: Name is empty in {file_path}.")

    # check that the number of prefs matches the number of trips, return for this
    if len(prefs) != numberOfTrips:
        print(
            f"Error: Number of preferences does not match number of trips in {file_path}. This leader will not be added."
        )
        return

    leader = TripLeader(name, prefs)
    trip_leader_manager.add_trip_leader(leader)


def addTrips(trip_manager, numberOfTrips, tripDF, dateXY, tripXY, file_path):
    # iterate through the rows of the tripDF, adding each trip date and name to the trip_manager
    currentRow = dateXY[0] + 1

    while currentRow < len(tripDF) and not pd.isnull(
        tripDF.iloc[currentRow, dateXY[1]]
    ):
        name = tripDF.iloc[currentRow, tripXY[1]]
        date = tripDF.iloc[currentRow, dateXY[1]]
        trip_manager.add_trip(name, date)
        currentRow += 1

    # return if that the number of trips added matches the number of trips in the tripDF

    if len(trip_manager.get_trips()) != numberOfTrips:
        print(
            f"Error: Number of trips added does not match number of trips in {file_path}. Expected {numberOfTrips}, got {len(trip_manager.get_trips())}."
        )
        return


def process_excel_files(
    trip_leader_manager,
    trip_manager,
    numberOfTrips,
    dateXY,
    tripXY,
    prefXY,
    nameXY,
    prefsSheetIndex,
    tripLeaderInfoIndex,
    folder_path="Data",
):
    if not os.path.exists(folder_path):
        print("Folder does not exist.")
        return

    files = [
        f
        for f in os.listdir(folder_path)
        if f.endswith(".xlsx") and not f.startswith("~")
    ]

    if not files:
        print("No valid Excel files found in the folder.")
        return

    for index, file_name in enumerate(files):
        file_path = os.path.join(folder_path, file_name)

        # Checking if the file is empty
        try:
            df = pd.ExcelFile(file_path, engine="openpyxl")
            sheetNames = df.sheet_names
            prefsDF = pd.read_excel(file_path, sheet_name=sheetNames[prefsSheetIndex])
            tripLeaderDF = pd.read_excel(
                file_path, sheet_name=sheetNames[tripLeaderInfoIndex]
            )

            if prefsDF.empty or tripLeaderDF.empty:
                print(f"File {file_name} has an empty sheet, skipping.")
                continue
        except Exception as e:
            print(f"Could not read {file_name}: {e}")
            continue

        if index == 0:

            addTrips(trip_manager, numberOfTrips, prefsDF, dateXY, tripXY, file_path)

            createLeader(
                prefsDF,
                tripLeaderDF,
                numberOfTrips,
                trip_leader_manager,
                nameXY,
                prefXY,
                tripXY,
                file_path,
            )

        else:
            createLeader(
                prefsDF,
                tripLeaderDF,
                numberOfTrips,
                trip_leader_manager,
                nameXY,
                prefXY,
                tripXY,
                file_path,
            )


def createExcelFile(trip_leader_manager, trip_manager):
    outputFileName = "output.xlsx"

    # if output file already exists, delete it
    try:
        if os.path.exists(outputFileName):
            os.remove(outputFileName)
    except Exception as e:
        print("Error: Cannot have the file open. Details:", e)

    # create an empty dataframe with "Dates" and "TRiP" as columns
    df = pd.DataFrame(columns=["Dates", "TRiP"])

    # populate the first and second columns with trip dates and names
    trip_data = [
        {"Dates": trip.date, "TRiP": trip.name} for trip in trip_manager.get_trips()
    ]
    df = pd.concat([df, pd.DataFrame(trip_data)], ignore_index=True)

    # populate the rest of the columns with the header of the trip leader name, and underneath their preferences
    for leader in trip_leader_manager.get_all_trip_leaders():
        # Ensure the list of preferences is the same length as the number of trips
        prefs = leader.prefs + [None] * (len(df) - len(leader.prefs))
        df[leader.name] = prefs

    # write the dataframe to an excel file
    df.to_excel(outputFileName, index=False)

    # Load the workbook and select the active worksheet
    wb = load_workbook(outputFileName)
    ws = wb.active

    # Bold the headers and center all cells
    header_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")

    # Apply formatting to headers
    for cell in ws[1]:  # First row (headers)
        cell.font = header_font
        cell.alignment = center_alignment

    # Colors for different categories
    color_map = {
        "bottom third": PatternFill(
            start_color="00FF00", end_color="00FF00", fill_type="solid"
        ),  # Red
        "middle third": PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
        ),  # Yellow
        "top third": PatternFill(
            start_color="FF0000", end_color="FF0000", fill_type="solid"
        ),  # Green
        "nan": PatternFill(
            start_color="000000", end_color="000000", fill_type="solid"
        ),  # Black
    }

    # Apply center alignment and color based on preference category
    for i, leader in enumerate(
        trip_leader_manager.get_all_trip_leaders(), start=3
    ):  # Columns start from C
        categories = leader.categorize_prefs()
        for row_num, (pref, category) in enumerate(
            categories, start=2
        ):  # Rows start from 2 (first row is header)
            cell = ws.cell(row=row_num, column=i)
            cell.alignment = center_alignment

            if pd.isna(pref):
                cell.fill = color_map["nan"]
            else:
                cell.fill = color_map.get(category, None)

    # Save the formatted Excel file
    wb.save(outputFileName)

    print(f"Excel file '{outputFileName}' created and formatted successfully.")
